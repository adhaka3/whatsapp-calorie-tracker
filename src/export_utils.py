"""Utility functions for exporting meal data"""
import json
from typing import Optional, List, Dict


def export_to_excel(db, output_file: str = "meal_logs.xlsx", phone_number: Optional[str] = None):
    """Export meal logs to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import sqlite3

        conn = sqlite3.connect(db.db_path)
        cursor = conn.cursor()

        # Query all meals or for specific user
        if phone_number:
            cursor.execute('''
                SELECT phone_number, meal_description, timestamp,
                       items_extracted, total_calories, total_protein, source, meal_tag
                FROM meals
                WHERE phone_number = ?
                ORDER BY timestamp DESC
            ''', (phone_number,))
        else:
            cursor.execute('''
                SELECT phone_number, meal_description, timestamp,
                       items_extracted, total_calories, total_protein, source, meal_tag
                FROM meals
                ORDER BY timestamp DESC
            ''')
        
        rows = cursor.fetchall()
        conn.close()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Meal Logs"
        
        # Headers
        headers = ["Phone Number", "Original Message", "Timestamp", "Meal Tag",
                  "Items Extracted", "Total Calories (kcal)", "Total Protein (g)", "Source"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row in rows:
            phone, description, timestamp, items_extracted, calories, protein, source, meal_tag = row
            
            # Parse items_extracted if it's JSON
            if items_extracted:
                try:
                    items_list = json.loads(items_extracted) if items_extracted.startswith('[') else items_extracted
                    if isinstance(items_list, list):
                        items_str = ", ".join([f"{item.get('quantity', 1)}x {item.get('name', item.get('food', 'unknown'))}" 
                                              for item in items_list])
                    else:
                        items_str = items_extracted
                except:
                    items_str = items_extracted
            else:
                items_str = "N/A"

            # Format meal tag for display
            meal_tag_display = meal_tag.replace('_', ' ').title() if meal_tag else "N/A"

            ws.append([
                phone,
                description,
                timestamp,
                meal_tag_display,
                items_str,
                round(calories, 1) if calories else 0,
                round(protein, 1) if protein else 0,
                source or "whatsapp"
            ])

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12
        
        # Save file
        wb.save(output_file)
        return True, f"✅ Exported {len(rows)} meals to {output_file}"
        
    except Exception as e:
        return False, f"❌ Error exporting to Excel: {e}"
