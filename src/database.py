"""Database handler for user meal tracking"""
import sqlite3
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import os
import json


def get_meal_tag(timestamp: datetime = None) -> str:
    """
    Determine meal tag based on timestamp.

    Time ranges:
    - Breakfast: 5:00 AM - 10:59 AM
    - Brunch: 11:00 AM - 11:59 AM
    - Lunch: 12:00 PM - 2:59 PM
    - Evening Snack: 3:00 PM - 5:59 PM
    - Dinner: 6:00 PM - 9:59 PM
    - Midnight Snack: 10:00 PM - 4:59 AM
    """
    if timestamp is None:
        timestamp = datetime.now()

    # Handle string timestamps
    if isinstance(timestamp, str):
        try:
            timestamp = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
        except:
            timestamp = datetime.now()

    hour = timestamp.hour

    if 5 <= hour < 11:
        return "breakfast"
    elif 11 <= hour < 12:
        return "brunch"
    elif 12 <= hour < 15:
        return "lunch"
    elif 15 <= hour < 18:
        return "evening_snack"
    elif 18 <= hour < 22:
        return "dinner"
    else:  # 22-4 (10 PM - 4:59 AM)
        return "midnight_snack"


class MealDatabase:
    def __init__(self, db_path: str = "data/user_meals.db"):
        self.db_path = db_path
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        self.init_database()
    
    def init_database(self):
        """Initialize the database with required tables"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Create users table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                phone_number TEXT PRIMARY KEY,
                name TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create meals table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS meals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                phone_number TEXT,
                meal_description TEXT,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                total_calories REAL,
                total_protein REAL,
                parsed_items TEXT,
                items_extracted TEXT,
                source TEXT DEFAULT 'whatsapp',
                meal_tag TEXT,
                FOREIGN KEY (phone_number) REFERENCES users(phone_number)
            )
        ''')

        # Add meal_tag column if it doesn't exist (for existing databases)
        try:
            cursor.execute('ALTER TABLE meals ADD COLUMN meal_tag TEXT')
            conn.commit()
        except sqlite3.OperationalError:
            # Column already exists
            pass

        # Create custom_foods table for user-added foods
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS custom_foods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                aliases TEXT,
                calories REAL NOT NULL,
                protein REAL NOT NULL,
                serving_size TEXT NOT NULL,
                category TEXT DEFAULT 'custom',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        conn.commit()
        conn.close()
    
    def add_user(self, phone_number: str, name: Optional[str] = None):
        """Add a new user to the database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT OR IGNORE INTO users (phone_number, name)
            VALUES (?, ?)
        ''', (phone_number, name))
        
        conn.commit()
        conn.close()
    
    def log_meal(self, phone_number: str, meal_description: str,
                 total_calories: float, total_protein: float,
                 parsed_items: str, items_extracted: str = "",
                 source: str = "whatsapp", timestamp: datetime = None):
        """Log a meal for a user"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Ensure user exists
        self.add_user(phone_number)

        # Determine meal tag based on timestamp
        if timestamp is None:
            timestamp = datetime.now()
        meal_tag = get_meal_tag(timestamp)

        # Convert timestamp to ISO format string
        timestamp_str = timestamp.isoformat() if isinstance(timestamp, datetime) else timestamp

        cursor.execute('''
            INSERT INTO meals (phone_number, meal_description, timestamp, total_calories,
                             total_protein, parsed_items, items_extracted, source, meal_tag)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (phone_number, meal_description, timestamp_str, total_calories,
              total_protein, parsed_items, items_extracted, source, meal_tag))

        conn.commit()
        conn.close()
    
    def get_daily_summary(self, phone_number: str, date: Optional[datetime] = None) -> Dict:
        """Get daily summary of calories and protein for a user"""
        if date is None:
            date = datetime.now()
        
        date_str = date.strftime('%Y-%m-%d')
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT 
                COUNT(*) as meal_count,
                SUM(total_calories) as total_calories,
                SUM(total_protein) as total_protein
            FROM meals
            WHERE phone_number = ?
            AND DATE(timestamp) = ?
        ''', (phone_number, date_str))
        
        result = cursor.fetchone()
        conn.close()
        
        return {
            'date': date_str,
            'meal_count': result[0] or 0,
            'total_calories': round(result[1] or 0, 1),
            'total_protein': round(result[2] or 0, 1)
        }
    
    def get_recent_meals(self, phone_number: str, limit: int = 5) -> List[Dict]:
        """Get recent meals for a user"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT meal_description, timestamp, total_calories, total_protein, meal_tag
            FROM meals
            WHERE phone_number = ?
            ORDER BY timestamp DESC
            LIMIT ?
        ''', (phone_number, limit))

        meals = []
        for row in cursor.fetchall():
            meals.append({
                'description': row[0],
                'timestamp': row[1],
                'calories': round(row[2], 1),
                'protein': round(row[3], 1),
                'meal_tag': row[4]
            })

        conn.close()
        return meals

    def delete_last_meal(self, phone_number: str) -> Dict:
        """
        Delete the most recent meal for a user

        Returns:
            Dictionary with success status and deleted meal info
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Get the most recent meal
            cursor.execute('''
                SELECT id, meal_description, total_calories, total_protein, timestamp, meal_tag
                FROM meals
                WHERE phone_number = ?
                ORDER BY timestamp DESC
                LIMIT 1
            ''', (phone_number,))

            result = cursor.fetchone()

            if not result:
                conn.close()
                return {
                    'success': False,
                    'message': '❌ No meals found to delete.\n\n'
                              'You haven\'t logged any meals yet!'
                }

            meal_id, description, calories, protein, timestamp, meal_tag = result

            # Delete the meal
            cursor.execute('DELETE FROM meals WHERE id = ?', (meal_id,))
            conn.commit()

            # Format meal tag for display
            meal_tag_display = meal_tag.replace('_', ' ').title() if meal_tag else "N/A"

            # Parse timestamp
            try:
                if isinstance(timestamp, str):
                    ts = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                    time_str = ts.strftime('%I:%M %p')
                else:
                    time_str = "Unknown time"
            except:
                time_str = str(timestamp)

            conn.close()

            return {
                'success': True,
                'message': f'✅ *Last Meal Deleted*\n\n'
                          f'🗑️ Removed: {description[:50]}\n'
                          f'🕒 Logged at: {time_str}\n'
                          f'🏷️ Meal Tag: {meal_tag_display}\n'
                          f'🔥 Calories: {round(calories, 1)} kcal\n'
                          f'💪 Protein: {round(protein, 1)}g\n\n'
                          f'Your daily totals have been updated.',
                'deleted_meal': {
                    'description': description,
                    'calories': round(calories, 1),
                    'protein': round(protein, 1),
                    'timestamp': timestamp,
                    'meal_tag': meal_tag
                }
            }

        except Exception as e:
            conn.close()
            return {
                'success': False,
                'message': f'❌ Error deleting meal: {str(e)}'
            }
    
    def get_weekly_summary(self, phone_number: str) -> Dict:
        """Get weekly summary"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT
                COUNT(*) as meal_count,
                AVG(total_calories) as avg_calories,
                AVG(total_protein) as avg_protein,
                SUM(total_calories) as total_calories,
                SUM(total_protein) as total_protein
            FROM meals
            WHERE phone_number = ?
            AND timestamp BETWEEN ? AND ?
        ''', (phone_number, start_date, end_date))

        result = cursor.fetchone()
        conn.close()

        return {
            'period': f'Last 7 days',
            'meal_count': result[0] or 0,
            'avg_daily_calories': round(result[1] or 0, 1),
            'avg_daily_protein': round(result[2] or 0, 1),
            'total_calories': round(result[3] or 0, 1),
            'total_protein': round(result[4] or 0, 1)
        }

    def get_weekly_breakdown(self, phone_number: str) -> Dict:
        """
        Get daily breakdown for the last 7 days

        Returns:
            Dictionary with daily breakdown and totals
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Get data for last 7 days
        daily_data = []
        total_calories = 0
        total_protein = 0
        total_meals = 0

        for i in range(6, -1, -1):  # 6 days ago to today
            date = datetime.now() - timedelta(days=i)
            date_str = date.strftime('%Y-%m-%d')

            cursor.execute('''
                SELECT
                    COUNT(*) as meal_count,
                    SUM(total_calories) as total_calories,
                    SUM(total_protein) as total_protein
                FROM meals
                WHERE phone_number = ?
                AND DATE(timestamp) = ?
            ''', (phone_number, date_str))

            result = cursor.fetchone()
            meal_count = result[0] or 0
            calories = result[1] or 0
            protein = result[2] or 0

            # Format day name
            if i == 0:
                day_label = "Today"
            elif i == 1:
                day_label = "Yesterday"
            else:
                day_label = date.strftime('%A')  # Full day name (Monday, Tuesday, etc.)

            daily_data.append({
                'date': date_str,
                'day_label': day_label,
                'day_name': date.strftime('%a'),  # Short day name (Mon, Tue, etc.)
                'full_date': date.strftime('%b %d'),  # Month Day (Jan 15)
                'meal_count': meal_count,
                'calories': round(calories, 1),
                'protein': round(protein, 1)
            })

            total_calories += calories
            total_protein += protein
            total_meals += meal_count

        conn.close()

        # Calculate averages (only for days with meals)
        days_with_meals = sum(1 for day in daily_data if day['meal_count'] > 0)
        avg_calories = (total_calories / days_with_meals) if days_with_meals > 0 else 0
        avg_protein = (total_protein / days_with_meals) if days_with_meals > 0 else 0

        return {
            'daily_breakdown': daily_data,
            'total_calories': round(total_calories, 1),
            'total_protein': round(total_protein, 1),
            'total_meals': total_meals,
            'avg_daily_calories': round(avg_calories, 1),
            'avg_daily_protein': round(avg_protein, 1),
            'days_with_meals': days_with_meals
        }
    
    def export_to_excel(self, output_file: str = "meal_logs.xlsx", phone_number: Optional[str] = None):
        """Export meal logs to Excel file"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Query all meals or for specific user
            if phone_number:
                cursor.execute('''
                    SELECT phone_number, meal_description, timestamp,
                           items_extracted, total_calories, total_protein, source, meal_tag
                    FROM meals
                    WHERE phone_number = ?
                    ORDER BY timestamp DESC
                ''', (phone_number,))
            else:
                cursor.execute('''
                    SELECT phone_number, meal_description, timestamp,
                           items_extracted, total_calories, total_protein, source, meal_tag
                    FROM meals
                    ORDER BY timestamp DESC
                ''')

            rows = cursor.fetchall()
            conn.close()

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Meal Logs"

            # Headers
            headers = ["Phone Number", "Original Message", "Timestamp", "Meal Tag",
                      "Items Extracted", "Total Calories", "Total Protein", "Source"]
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data rows
            for row in rows:
                phone, description, timestamp, items_extracted, calories, protein, source, meal_tag = row

                # Parse items_extracted if it's JSON
                if items_extracted:
                    try:
                        items_list = json.loads(items_extracted) if items_extracted.startswith('[') else items_extracted
                        if isinstance(items_list, list):
                            items_str = ", ".join([f"{item.get('quantity', 1)}x {item.get('name', item.get('food', 'unknown'))}"
                                                  for item in items_list])
                        else:
                            items_str = items_extracted
                    except:
                        items_str = items_extracted
                else:
                    items_str = "N/A"

                # Format meal tag for display
                meal_tag_display = meal_tag.replace('_', ' ').title() if meal_tag else "N/A"

                ws.append([
                    phone,
                    description,
                    timestamp,
                    meal_tag_display,
                    items_str,
                    round(calories, 1) if calories else 0,
                    round(protein, 1) if protein else 0,
                    source or "unknown"
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12

            # Save file
            wb.save(output_file)
            return True, f"Exported {len(rows)} meals to {output_file}"

        except Exception as e:
            return False, f"Error exporting to Excel: {e}"
    
    def get_all_meals(self, phone_number: Optional[str] = None, limit: Optional[int] = None) -> List[Dict]:
        """Get all meals for export or analysis"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        if phone_number:
            if limit:
                cursor.execute('''
                    SELECT meal_description, timestamp, items_extracted,
                           total_calories, total_protein, source, meal_tag
                    FROM meals
                    WHERE phone_number = ?
                    ORDER BY timestamp DESC
                    LIMIT ?
                ''', (phone_number, limit))
            else:
                cursor.execute('''
                    SELECT meal_description, timestamp, items_extracted,
                           total_calories, total_protein, source, meal_tag
                    FROM meals
                    WHERE phone_number = ?
                    ORDER BY timestamp DESC
                ''', (phone_number,))
        else:
            if limit:
                cursor.execute('''
                    SELECT meal_description, timestamp, items_extracted,
                           total_calories, total_protein, source, meal_tag
                    FROM meals
                    ORDER BY timestamp DESC
                    LIMIT ?
                ''', (limit,))
            else:
                cursor.execute('''
                    SELECT meal_description, timestamp, items_extracted,
                           total_calories, total_protein, source, meal_tag
                    FROM meals
                    ORDER BY timestamp DESC
                ''')

        meals = []
        for row in cursor.fetchall():
            meals.append({
                'description': row[0],
                'timestamp': row[1],
                'items_extracted': row[2],
                'calories': round(row[3], 1),
                'protein': round(row[4], 1),
                'source': row[5],
                'meal_tag': row[6]
            })

        conn.close()
        return meals

    def add_custom_food(self, name: str, calories: float, protein: float, serving_size: str, category: str = "custom") -> Dict:
        """
        Add a custom food to the database (shared by all users)

        Args:
            name: Food name (e.g., "protein shake")
            calories: Calories per serving
            protein: Protein in grams per serving
            serving_size: Description of serving size (e.g., "1 scoop (30g)")
            category: Optional category (default: "custom")

        Returns:
            Dictionary with status and message
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Check if food already exists
            cursor.execute('SELECT name, calories, protein, serving_size FROM custom_foods WHERE name = ?', (name.lower(),))
            existing = cursor.fetchone()

            if existing:
                conn.close()
                return {
                    'success': False,
                    'message': f'❌ Food "{name}" already exists in database.\n\n'
                              f'Current values:\n'
                              f'  Calories: {existing[1]} kcal\n'
                              f'  Protein: {existing[2]}g\n'
                              f'  Serving: {existing[3]}'
                }

            # Insert new custom food
            cursor.execute('''
                INSERT INTO custom_foods (name, aliases, calories, protein, serving_size, category)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (name.lower(), '[]', float(calories), float(protein), serving_size, category))

            conn.commit()
            conn.close()

            return {
                'success': True,
                'message': f'✅ Food "{name}" added successfully!'
            }

        except sqlite3.IntegrityError:
            return {
                'success': False,
                'message': f'❌ Food "{name}" already exists in database.'
            }
        except Exception as e:
            return {
                'success': False,
                'message': f'❌ Error adding food: {str(e)}'
            }

    def get_all_custom_foods(self) -> List[Dict]:
        """
        Get all custom foods from the database

        Returns:
            List of custom food dictionaries
        """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('''
            SELECT name, aliases, calories, protein, serving_size, category
            FROM custom_foods
            ORDER BY name
        ''')

        custom_foods = []
        for row in cursor.fetchall():
            custom_foods.append({
                'name': row[0],
                'aliases': json.loads(row[1]) if row[1] else [],
                'calories': row[2],
                'protein': row[3],
                'serving_size': row[4],
                'category': row[5] if row[5] else 'custom'
            })

        conn.close()
        return custom_foods

    def delete_custom_food(self, name: str) -> Dict:
        """
        Delete a custom food from the database

        Args:
            name: Food name to delete

        Returns:
            Dictionary with status and message
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            cursor.execute('DELETE FROM custom_foods WHERE name = ?', (name.lower(),))

            if cursor.rowcount == 0:
                conn.close()
                return {
                    'success': False,
                    'message': f'❌ Food "{name}" not found in custom foods.'
                }

            conn.commit()
            conn.close()

            return {
                'success': True,
                'message': f'✅ Food "{name}" deleted successfully!'
            }

        except Exception as e:
            return {
                'success': False,
                'message': f'❌ Error deleting food: {str(e)}'
            }